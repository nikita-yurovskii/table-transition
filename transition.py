import docx
from docx.shared import Pt
from docx.enum.text import WD_ALIGN_PARAGRAPH
import openpyxl
from openpyxl.utils import range_boundaries
from openpyxl.utils import range_to_tuple, coordinate_to_tuple
import excel_table
def transfer_excel_to_word_table(file, sheet, word_table,tpp):
    """
    Transfers data from an Excel range to a Word table.
    Sets font to Times New Roman, 8pt, and aligns text.

    Args:
        word_table (docx.table.Table): The Word table object to populate.
        excel_range (openpyxl.worksheet.worksheet.Worksheet): The Excel range to read from (e.g., ws['A1:C5']).
    """

    min_row, max_row , min_col, max_col  = tpp

    # Get the number of rows and columns in the Excel range
    excel_rows = max_row - min_row + 1
    excel_cols = max_col - min_col + 1

    # Get the number of rows and columns in the Word table
    word_rows = len(word_table.rows)
    word_cols = len(word_table.columns)

    # Resize Word table if necessary (add rows and columns).
    # Add extra rows

    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet]
    # Transfer data from Excel to Word
    for i in range(excel_rows):
        for j in range(excel_cols):
            excel_cell = sheet.cell(min_row + i, min_col + j)
            word_cell = word_table.cell(i, j)  # Word table indices are 0-based

            # Get the cell value from Excel

            cell_value = excel_cell.value
            print(i,j,cell_value)
            if cell_value:
                word_cell.text = str(cell_value)
            else:
                word_cell.text = ''
            run = word_cell.paragraphs[0].runs[0]
            run.font.name = 'Times New Roman'
            run.font.size = Pt(8)
            print(word_cell.text)
            # if word_cell.text:
            #     word_cell.text = cell_value
            # Add a paragraph to the Word cell with the Excel value



            # Set font and alignment
        #    for run in paragraph.runs: # Important, apply formatting to each run (section)
        #        run.font.name = 'Times New Roman'
        #        run.font.size = Pt(8)
        #    paragraph.alignment = WD_ALIGN_PARAGRAPH.CENTER  # Example alignment



def get_excel_range(sheet, start_row, start_col, end_row, end_col):
    """
    Gets an Excel range from a worksheet using row and column coordinates.

    Args:
        sheet (openpyxl.worksheet.worksheet.Worksheet): The Excel worksheet object.
        start_row (int): The starting row number (1-based).
        start_col (int): The starting column number (1-based).
        end_row (int): The ending row number (1-based).
        end_col (int): The ending column number (1-based).

    Returns:
        openpyxl.worksheet.worksheet.Worksheet: The Excel range as openpyxl.worksheet.worksheet.Worksheet object.
        Note: This *returns* the same sheet object, as openpyxl doesn't natively have a range object.
        The returned object can be used as the excel_range parameter for `transfer_excel_to_word_table`.
        The user should use the sheet, and specify the range with openpyxl.utils.range_string.
    """

    range_address = coordinate_to_tuple(start_row, start_col, end_row, end_col)
    return sheet[range_address]
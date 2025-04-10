import openpyxl
from docx import Document
from docx.enum.table import WD_ALIGN_VERTICAL
from docx.oxml.table import CT_Tbl
from docx.oxml.text.paragraph import CT_P
from docx.table import Table, _Cell

def get_merged_cells(file_path, pogr_r,pogr_c,mr,mc, sheet_name=None):
    """
    Gets a list of merged cells from an Excel file in a nested array format.

    Args:
        file_path (str): The path to the Excel file.
        sheet_name (str, optional): The name of the sheet. If None, the active sheet is used. Defaults to None.

    Returns:
        list: A list of nested arrays, where each sub-array represents a merged cell range.
              For example: [['A1', 'B2'], ['C3', 'D4']]
              Returns an empty list if there are no merged cells.
    """
    print('eto_pogr_c:',pogr_c, pogr_r)
    wb = openpyxl.load_workbook(file_path)
    a = []
    if sheet_name:
        try:
            sheet = wb[sheet_name]
        except KeyError:
            print(f"Sheet '{sheet_name}' not found.")
            return []
    else:
        sheet = wb.active

    print('----------- ', sheet.merged_cells.ranges)
    for i in sheet.merged_cells.ranges:
        a.append(list(i.cells))

    for i in a:
        for j in range(len(i)):
            i[j] = list(i[j])
            i[j][0]-=pogr_r+1
            i[j][1]-=(pogr_c)


            if i[j][0] > mr or i[j][1] > mc or i[j][0] < 0 or i[j][1]<0:
                print("\033[96m coords", i[j][0], i[j][1],"were too shitty so \033[0m")
                for k in range(len(i)):
                    i[j] = list(i[j])
                    i[j][0] = 0
                    i[j][1] = 0
    return a




def merge_cells_from_coordinates(document, table, coordinates):
    print(table.cell)
    """
    Объединяет ячейки в таблице Word по заданным координатам.

    Args:
        document: Объект документа Word (docx.Document) - для работы с таблицами.
        table: Объект таблицы (docx.table.Table).
        coordinates: Список кортежей, представляющих координаты ячеек для объединения.
                     Каждый кортеж имеет вид (row_index, col_index).
                     Координаты предполагаются последовательными и образовывающими
                     прямоугольную область.
    """
    to_merge = []
    for i in coordinates:
        print('started merging for', table, i)
        if not i:
            print("Предупреждение: Пустой список координат. Объединение не выполнено.")
            return
        print(i, 'coords')
        # 1. Найти минимальные и максимальные значения row и col
        min_row = min(coord[0] for coord in i)
        max_row = max(coord[0] for coord in i)
        min_col = min(coord[1] for coord in i)
        max_col = max(coord[1] for coord in i)

        # 2. Валидация координат (проверка, что координаты образуют прямоугольник)
        expected_count = (max_row - min_row + 1) * (max_col - min_col + 1)
        if len(i) != expected_count:
            print("Ошибка: Координаты",min_row, min_col, max_row, max_col," не образуют прямоугольник.  Объединение не выполнено.")
            continue

        # 3. Объединение ячеек

        first_cell = table.cell(min_row, min_col)
        last_cell = table.cell(max_row, max_col)
        print('new first and last cell-', (min_row, min_col), (max_row, max_col))
        to_merge.append([first_cell,last_cell, (min_row,min_col), (max_row,max_col)])

        print("merged", first_cell, last_cell)

        # Удаляем содержимое из объединяемых ячеек (кроме первой)
    k = 0
    for i in to_merge:
        min_row, min_col = i[2][0],i[2][1]
        max_row, max_col = i[3][0],i[3][1]
        first_cell = i[0]
        last_cell = i[1]
        merged_text = first_cell.text  # Сохраняем текст из первой ячейки
        run = first_cell.paragraphs[0].runs[0]
        mn = run.font.name
        ms = run.font.size
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = table.cell(row, col)
                if row != min_row or col != min_col: # Не удаляем текст из первой ячейки
                    cell.text = ""

        if min_row >=0 and min_col>=0 and max_col>=0 and max_row>=0:
            print('\033[92m Trying to merge \033[0m', i[2], i[3], first_cell, last_cell)
            try:
                first_cell.merge(last_cell)  # Объединяем ячейки
                first_cell.text = merged_text # Возвращаем текст в объединенную ячейку
                run = first_cell.paragraphs[0].runs[0]
                run.font.name = mn
                run.font.size = ms
                document.save("govno/output"+str(i[2]) +str(i[3])+".docx")
                k+=1
            except Exception:
                print('doh')




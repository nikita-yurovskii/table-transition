import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from termcolor import colored



def get_print_area(file_path, sheet_name):
    """
        Получает размер области печати Excel (строки, столбцы).
        Возвращает (None, None) если нет области печати или произошла ошибка.

        Args:
            file_path (str): Путь к файлу Excel.
            sheet_name (str): Имя листа, с которым нужно работать.

        Returns:
            tuple: Кортеж, содержащий:
                - Количество строк в области печати.
                - Количество столбцов в области печати.
                Возвращает (None, None) если область печати не задана или произошла ошибка.
                Возвращает (None, None) в случае ошибки при чтении файла.
        """
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]

        print_area = sheet.print_area

        if not print_area:
            return None, None  # Область печати не задана

        # Очищаем от лишних пробелов и приводим к верхнему регистру
        print_area = print_area.strip().upper()
        print('eto pa:', print_area)
        # Разбираем print_area, который представляет собой строку, например, "A1:C10"
        try:
            pa = print_area.split("!")[1]
            start_cell, end_cell = pa.split(':')
            print(start_cell, end_cell, 'se')
        except ValueError:
            print(f"Некорректный формат области печати: {print_area}")
            return None, None

        def extract_col_row(cell):
            """Извлекает букву столбца и номер строки из ячейки, обрабатывая ошибки"""
            col_letter = ''.join([char for char in cell if char.isalpha()])
            row_str = ''.join([char for char in cell if char.isdigit()])
            row = int(row_str) if row_str else None  # Обработка случая, когда строка отсутствует
            return col_letter, row

        start_col_letter, start_row = extract_col_row(start_cell)
        end_col_letter, end_row = extract_col_row(end_cell)

        # Преобразование буквенных обозначений столбцов в индексы
        try:
            start_col_idx = openpyxl.utils.column_index_from_string(start_col_letter)
            end_col_idx = openpyxl.utils.column_index_from_string(end_col_letter)
        except ValueError as e:
            print(f"Некорректное имя столбца в области печати: {e}")
            return None, None

        # Обработка случаев, когда строка не указана
        if start_row is None or end_row is None:
            return None, None  # Возвращаем None, None если не можем определить размер
        return start_row,end_row,start_col_idx,end_col_idx
    except Exception as e:
        print(f"Ошибка при чтении файла Excel: {e}")
        return None, None

def create_worksheet_from_range(input_excel_file, worksheet_name):
    """
    Создает новый воркшит в новом Excel файле, копируя данные из указанного диапазона существующего Excel файла.

    Args:
        input_excel_file: Путь к существующему Excel файлу.
        output_excel_file: Путь к новому Excel файлу, который будет создан.
        worksheet_name: Имя нового воркшита.
        start_row: Номер первой строки диапазона (начиная с 1).
        end_row: Номер последней строки диапазона (начиная с 1).
        start_col_idx: Индекс первой колонки диапазона (начиная с 1).
        end_col_idx: Индекс последней колонки диапазона (начиная с 1).
    """

    try:

        start_row, end_row, start_col_idx, end_col_idx = get_print_area(input_excel_file, worksheet_name)
        # Загрузка существующего воркбука
        workbook = openpyxl.load_workbook(input_excel_file)

        # Получение активного воркшита (или можно указать имя воркшита, если нужно)
        source_sheet = workbook.active  #workbook["Sheet1"]

        # Создание нового воркбука
        new_workbook = openpyxl.Workbook()

        # Создание нового воркшита в новом воркбуке
        new_sheet = new_workbook.create_sheet(worksheet_name)


        # Копирование данных из диапазона
        for row_idx in range(start_row, end_row + 1):
            for col_idx in range(start_col_idx, end_col_idx + 1):
                cell_value = source_sheet.cell(row=row_idx, column=col_idx).value  # получение значения ячейки
                new_sheet.cell(row=row_idx - start_row + 1, column=col_idx - start_col_idx + 1).value = cell_value  # запись в новый воркшит
                # Copy cell styles
                source_cell = source_sheet.cell(row=row_idx, column=col_idx)
                target_cell = new_sheet.cell(row=row_idx - start_row + 1, column=col_idx - start_col_idx + 1)
                if source_cell.has_style:
                    target_cell.font = openpyxl.styles.Font(name=source_cell.font.name, size=source_cell.font.size, bold=source_cell.font.bold, italic=source_cell.font.italic, color=source_cell.font.color)
                    target_cell.border = openpyxl.styles.Border(left=source_cell.border.left, right=source_cell.border.right, top=source_cell.border.top, bottom=source_cell.border.bottom)
                    target_cell.fill = openpyxl.styles.PatternFill(fill_type=source_cell.fill.fill_type, fgColor=source_cell.fill.fgColor, bgColor=source_cell.fill.bgColor)
                    target_cell.number_format = source_cell.number_format
                    #target_cell.protection = source_cell.protection
                    target_cell.alignment = openpyxl.styles.Alignment(horizontal=source_cell.alignment.horizontal, vertical=source_cell.alignment.vertical, text_rotation=source_cell.alignment.text_rotation, wrap_text=source_cell.alignment.wrap_text, shrink_to_fit=source_cell.alignment.shrink_to_fit, indent=source_cell.alignment.indent)


        # Сохранение нового воркбука
        return new_sheet



    except FileNotFoundError:
        print(f"Ошибка: Файл '{input_excel_file}' не найден.")
    #except Exception as e:
       # print(f"Произошла ошибка: {e}")
def get_print_area_size(file_path, sheet_name):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
        start_row, end_row, start_col_idx, end_col_idx = get_print_area(file_path, sheet_name)

        num_rows = end_row - start_row + 1

        num_cols = end_col_idx - start_col_idx + 1

        #if num_rows < 0 or num_cols < 0:  # Проверка на некорректные значения. Убрано
        #    return 0, 0  # пустая область печати. Убрано
        print('\033[92m',start_col_idx,end_col_idx, '\033[0m')

        for col_idx in range(start_col_idx, end_col_idx + 1):
            col_letter = get_column_letter(col_idx)
            print('tried the', col_letter, sheet.column_dimensions[col_letter].hidden)
            if col_letter in sheet.column_dimensions and sheet.column_dimensions[col_letter].hidden:
                print('deleted the', col_letter)
                num_cols-=1


        return num_rows, num_cols




def get_first_visible_row_index(file_path, sheet_name):
    """
    Возвращает индекс первой не скрытой строки в листе Excel.

    Args:
        file_path (str): Путь к файлу Excel.
        sheet_name (str): Имя листа.

    Returns:
        int: Индекс первой не скрытой строки.
             Возвращает None, если все строки скрыты или произошла ошибка.
    """
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]

        for row_idx in range(1, sheet.max_row + 1):
            if row_idx in sheet.row_dimensions and not sheet.row_dimensions[row_idx].hidden:
                return row_idx
            elif row_idx not in sheet.row_dimensions:
                return row_idx  # Если строка не имеет настроек, считаем ее видимой.
        return None  # Все строки скрыты
    except Exception as e:
        print(f"Ошибка при чтении файла Excel: {e}")
        return None



def get_first_visible_column_index(file_path, sheet_name):
    """
    Возвращает индекс первого не скрытого столбца в листе Excel.

    Args:
        file_path (str): Путь к файлу Excel.
        sheet_name (str): Имя листа.

    Returns:
        int: Индекс первого не скрытого столбца.
             Возвращает None, если все столбцы скрыты или произошла ошибка.
    """
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]

        for col_idx in range(1, sheet.max_column + 1):
            col_letter = get_column_letter(col_idx)
            if col_letter in sheet.column_dimensions:
                if not sheet.column_dimensions[col_letter].hidden:
                    return col_idx
            else:
                return col_idx  # Если столбец не имеет настроек, считаем его видимым

        return None  # Все столбцы скрыты
    except Exception as e:
        print(f"Ошибка при чтении файла Excel: {e}")
        return None



def get_cell_widths_in_range(file_path, sheet_name, pogr_r, pogr_c):
    """
    Получает ширину каждой ячейки в указанном диапазоне Excel.

    Args:
        file_path (str): Путь к файлу Excel.
        sheet_name (str): Имя листа.
        start_row (int): Индекс первой строки диапазона (начиная с 1).
        start_col (int): Индекс первого столбца диапазона (начиная с 1).
        end_row (int): Индекс последней строки диапазона.
        end_col (int): Индекс последнего столбца диапазона.

    Returns:
        dict: Словарь, где ключ - кортеж (row_idx, col_idx), а значение - ширина ячейки.
              Возвращает None, если произошла ошибка.
    """
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
        start_row, end_row, start_col, end_col = get_print_area(file_path, sheet_name)
        cell_widths = {}
        cell_heights = {}
        for row_idx in range(start_row, end_row + 1):
            for col_idx in range(start_col, end_col + 1):
                col_letter = get_column_letter(col_idx)

                # Получаем ширину столбца
                if col_letter in sheet.column_dimensions:
                    width = sheet.column_dimensions[col_letter].width
                if row_idx in sheet.row_dimensions:
                    height = sheet.row_dimensions[row_idx].height

                else:
                    width = None  # Если ширина столбца не задана, присваиваем None

                if row_idx-pogr_r > 0 and col_idx-pogr_c >= 0:
                    cell_widths[(row_idx-pogr_r-1, col_idx-1-pogr_c)] = width
                    cell_heights[(row_idx-pogr_r-1, col_idx-1-pogr_c)] = height
        return (cell_widths, cell_heights)

    except Exception as e:
        print(f"Ошибка при чтении файла Excel: {e}")
        return None